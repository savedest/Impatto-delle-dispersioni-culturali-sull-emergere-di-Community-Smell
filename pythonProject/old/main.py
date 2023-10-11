# import pandas as pd
#
# # Percorso al file CSV
# input_file_path = 'C:\\Users\\saver\\PycharmProjects\\pythonProject\\input.csv'
# output_file_path = 'C:\\Users\saver\\PycharmProjects\\pythonProject\\output.csv'
#
# # Carica il file CSV in un DataFrame
# df = pd.read_csv(input_file_path, delimiter=';')
#
# # Filtra le righe in base al valore nella colonna 'countries'
# df_filtered = df[~df['countries'].str.contains('None', case=False, na=True)]
#
# # Salva il DataFrame filtrato in un nuovo file CSV
# df_filtered.to_csv(output_file_path, index=False, sep=';', mode = 'w+')
#
# import pandas as pd
# from openpyxl import Workbook
#
# # Percorso al file CSV
# input_file_path = 'C:\\Users\\saver\\PycharmProjects\\pythonProject\\esempio.csv'
# output_excel_file_path = 'C:\\Users\\saver\\PycharmProjects\\pythonProject\\outputese.xlsx'
#
# # Carica il file CSV in un DataFrame
# df = pd.read_csv(input_file_path, delimiter=';')
#
# print(df['countries'])
# # Converti tutti i valori nella colonna 'countries' in stringhe
# df['countries'] = df['countries'].astype(str)
#
# # Separa i valori nella colonna 'countries' e conta il numero di 'null'
# df['Null_count'] = df['countries'].apply(lambda x: x.count('Null'))
# print(df['Null_count'])
# # Calcola la percentuale di 'Null' rispetto al numero totale di elementi nella colonna 'countries'
# df['Null_percentage'] = df['Null_count'] / df['countries'].apply(lambda x: len(x.split(',')))
#
# # Definisci i range percentuali
# percent_ranges = [(0, 0.25), (0.25, 0.5), (0.5, 0.75), (0.75, 1)]
#
# # Inizializza un foglio di lavoro Excel
# wb = Workbook()
#
# # Itera attraverso i range percentuali
# for start, end in percent_ranges:
#
#     sheet_name = f'{int(start * 100)}-{int(end * 100)}%'
#
#     # Filtra le righe in base alla percentuale di 'Null' nel range specificato
#     if(end<1):
#         filtered_rows = df[(df['Null_percentage'] >= start) &
#                        (df['Null_percentage'] < end)]
#     else:
#         filtered_rows = df[(df['Null_percentage'] >= start) &
#                            (df['Null_percentage'] <= end)]
#
#
#     # Crea un foglio di lavoro Excel per il range corrente
#     sheet = wb.create_sheet(title=sheet_name)
#
#
#
#     # Rimuovi le colonne di conteggio e percentuale aggiunte
#     filtered_rows = filtered_rows.drop(columns=['Null_count', 'Null_percentage'])
#
#
#     # Scrivi i dati nel foglio di lavoro
#     for row in filtered_rows.itertuples(index=False, name=None):
#         sheet.append(row)
#
#
#
# # Rimuovi il foglio di lavoro di default
# wb.remove(wb['Sheet'])
#
# # Salva il foglio di lavoro Excel
# wb.save(output_excel_file_path)



import pandas as pd
from openpyxl import Workbook

# Percorso al file CSV
input_file_path = 'C:\\Users\\saver\\PycharmProjects\\pythonProject\\input.csv'
output_excel_file_path = 'C:\\Users\\saver\\PycharmProjects\\pythonProject\\outputese.xlsx'

# Carica il file CSV in un DataFrame
df = pd.read_csv(input_file_path, delimiter=';')

# Converti tutti i valori nella colonna 'countries' in stringhe
df['countries'] = df['countries'].astype(str)

# Separa i valori nella colonna 'countries' e conta il numero di 'null'
df['Null_count'] = df['countries'].apply(lambda x: x.count('Null'))

# Calcola la percentuale di 'Null' rispetto al numero totale di elementi nella colonna 'countries'
df['Null_percentage'] = df['Null_count'] / df['countries'].apply(lambda x: len(x.split(',')))

# Definisci i range percentuali
percent_ranges = [(0, 0.25), (0.25, 0.5), (0.5, 0.75), (0.75, 1)]

# Inizializza un foglio di lavoro Excel
wb = Workbook()

# Itera attraverso i range percentuali
for start, end in percent_ranges:

    sheet_name = f'{int(start * 100)}-{int(end * 100)}%'

    # Filtra le righe in base alla percentuale di 'Null' nel range specificato
    if end < 1:
        filtered_rows = df[(df['Null_percentage'] >= start) &
                           (df['Null_percentage'] < end)]
    else:
        filtered_rows = df[(df['Null_percentage'] >= start) &
                           (df['Null_percentage'] <= end)]

    # Crea un foglio di lavoro Excel per il range corrente
    sheet = wb.create_sheet(title=sheet_name)

    # Scrivi l'header
    header = list(filtered_rows.columns)
    sheet.append(header)

    # Rimuovi le colonne di conteggio e percentuale aggiunte
    filtered_rows = filtered_rows.drop(columns=['Null_count', 'Null_percentage'])

    # Scrivi i dati nel foglio di lavoro
    for row in filtered_rows.itertuples(index=False, name=None):
        sheet.append(row)

# Itera attraverso i fogli di lavoro nel file Excel
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]

    # Ottieni il numero di righe
    num_rows = sheet.max_row

    print(f"Sheet '{sheet_name}' ha {num_rows-1} righe")
# Rimuovi il foglio di lavoro di default
wb.remove(wb['Sheet'])



# Salva il foglio di lavoro Excel
wb.save(output_excel_file_path)


